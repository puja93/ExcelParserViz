#!/usr/bin/env python
# coding: utf-8


from openpyxl import load_workbook
import sys
import re
import numpy as np
import pandas as pd
import operator
import time




#APPEND NULL AT FIRST & LAST ROW COLUMN
def append_null_limiter(df):
    start = time.time()
    temp = pd.DataFrame([pd.Series()], columns=df.index).T
    df = pd.concat([temp, df, temp],axis=1,ignore_index=True)

    temp = pd.DataFrame([pd.Series()], columns=df.columns)
    df = pd.concat([temp, df, temp],ignore_index=True)

    #print('append_null_limiter time is ',time.time()-start)
    return df


# RESTRUCTURE A LIST to ROW-COLUMN SHAPE
def list_to_df(list_data,col_size):

    start = time.time()
    ctr=0
    new_list=[]
    while ctr<len(list_data):
        new_list.append(list_data[ctr:ctr+col_size])
        ctr+=col_size
    df = pd.DataFrame.from_records(new_list)
    #print('list_to_df time is ',time.time()-start)

    return df



#GET NUMBER OF BORDER FROM A CELL
def count_cell_border(cell):
    start = time.time()
    dum = []
    dum.append(cell.border.top.style)
    dum.append(cell.border.bottom.style)
    dum.append(cell.border.right.style)
    dum.append(cell.border.left.style)


    #print('count_cell_border time is ',time.time()-start)

    return (dum.count('thin') if (dum.count('thin')>1) else None)




#THE BACKBONE OF THIS CODE, PARSE BORDER TO NUMPY ARRAY
def get_table_border(fname,sheetname):

    start = time.time()
    workbook = load_workbook(filename=fname,read_only=True)
    wb = workbook[sheetname]

    #Cell range
    #cells = wb.dimensions.split(':')
    #f_cell = cells[0]
    #l_cell = cells[1]
    #print("Sheet {} has range from {} to {}".format(sheetname, f_cell, l_cell))

    #Get row-col info
    row = wb.max_row
    col = wb.max_column

    #Store to a list temporarily
    col_list = []
    for row in wb.rows:
        for cell in row:
            if(cell.border!=None):
                temp = count_cell_border(cell)
            else:
                temp = None
            col_list.append(temp)

    dframe = list_to_df(col_list,col)
    dframe = dframe.astype('Int64')

    #print('get_table_border time is ',time.time()-start)

    return dframe



def get_border_by_row(df):
    start = time.time()
    val = []
    mark = True
    for row in df.index:
        for col in df.columns:
            cond = pd.isnull(df.iloc[row,col])

            if(cond ^ mark):
                if(cond):
                    val.append([row,col-1])
                else:
                    val.append([row,col])
            mark = cond

    #print('get_border_by_row time is ',time.time()-start)

    return val


def get_border_by_col(df):
    start = time.time()
    val = []
    mark = True
    for col in df.columns:
        for row in df.index:
            cond = pd.isnull(df.iloc[row,col])

            if(cond ^ mark):
                if(cond):
                    val.append([row-1,col])
                else:
                    val.append([row,col])
            mark = cond

    #print('get_border_by_col time is ',time.time()-start)
    return val



def find_edge_multitable(df):

    start = time.time()

    val_row = get_border_by_row(df)
    val_col = get_border_by_col(df)


    #print(val_row,end='')
    #print('\n\n')
    #print(val_col,end='')


    val_all = [value for value in val_row if value in val_col]
    #print(val_all)

    val_final=[]
    [val_final.append(x) for x in val_all if x not in val_final]
    #print(val_final,end='')



    #GET ROW
    tamp = sorted(val_final, key=operator.itemgetter(0))
    #print(len(tamp))


    #GET COLUMN
    temp = sorted(val_final, key=operator.itemgetter(1))
    #print(len(temp))



    lst = [item[0] for item in temp]
    res = []
    [res.append(x) for x in lst if x not in res]


    res_final = []
    for index in range(0,len(res),2):
        temp = res[index:index+2]
        res_final.append(min(temp))
        res_final.append(max(temp))
        #print(min(temp))
        #print(sorted(temp,reverse=True))
    #print(res_final)



    cond = False
    ces = []
    for r in res:
        dummy = []
        idx = [x[0] for x in enumerate(tamp) if r == x[1][0]]
        for ids in idx:
            dummy.append(tamp[ids])
        dum = sorted(dummy, key=operator.itemgetter(1),reverse=cond)
        cond = not cond
        ces.append(dum[0][1])

    #print(res_final)
    #print(ces)
    #print(len(ces),len(res_final))

    index = min(len(ces),len(res_final))

    if(index%2==0):
        pass
    else:
        index = index-1
    #print(index,len(ces),len(res_final))



    total_data = []
    for idx in range(0,index,2):
        val = (res_final[idx+1] - res_final[idx] + 1)*(ces[idx+1] - ces[idx]+1)
        total_data.append([val,idx,idx+1])

    final_data_index = sorted(total_data, key=operator.itemgetter(0), reverse=True)

    #print('find_edge_multitable time is ',time.time()-start)


    return res_final, ces, final_data_index
    #idx = total_data[0][1]
    #display(dataframe.iloc[res_final[idx]-1:res_final[idx+1],ces[idx]-1:ces[idx+1]+1])

def auto_parse(filename, sheet, debug=False):

    start = time.time()
    # Data Only
    #dfs = pd.read_excel(fname,sheet_name=None,header=None)
    #dataframe = dfs[sheetname]
    #print('read_excel time is ',time.time()-start)


    # Table Structure

    df = get_table_border(fname=filename, sheetname=sheet)

    #print('get_table_border is ',time.time()-start)

    #df.astype('Int32')
    df = append_null_limiter(df)
    #print('append_null_limiter is ',time.time()-start)

    if(debug):
        #print('Execution time is ', time.time()-start)
        print('Clustered Dataframe')
        display(df)
    #res_final, ces, all_idx = find_edge_multitable(df)
    #print(all_idx)
    #idx = all_idx[0][1]
    #print(idx)
    #display(dataframe.iloc[res_final[idx]-1:res_final[idx+1],ces[idx]-1:ces[idx+1]+1])

    #out = dataframe.iloc[res_final[idx]-1:res_final[idx+1],ces[idx]-1:ces[idx+1]+1]
    #out.to_csv('out.csv',encoding='utf-8-sig',index=False, header=False)

    #print('auto_parse time is ',time.time()-start)

    return df

