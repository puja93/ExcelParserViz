#!/usr/bin/env python
# coding: utf-8


from openpyxl import load_workbook
import sys
import re
import numpy as np
import pandas as pd
import operator
import time




#APPEND NULL AT FIRST & LAST ROW COLUMN
def append_null_limiter(df):
    start = time.time()
    temp = pd.DataFrame([pd.Series()], columns=df.index).T
    df = pd.concat([temp, df, temp],axis=1,ignore_index=True)

    temp = pd.DataFrame([pd.Series()], columns=df.columns)
    df = pd.concat([temp, df, temp],ignore_index=True)

    #print('append_null_limiter time is ',time.time()-start)
    return df


# RESTRUCTURE A LIST to ROW-COLUMN SHAPE
def list_to_df(list_data,col_size):

    start = time.time()
    ctr=0
    new_list=[]
    while ctr<len(list_data):
        new_list.append(list_data[ctr:ctr+col_size])
        ctr+=col_size
    df = pd.DataFrame.from_records(new_list)
    #print('list_to_df time is ',time.time()-start)

    return df



#GET NUMBER OF BORDER FROM A CELL
def count_cell_border(cell):
    start = time.time()
    dum = []
    dum.append(cell.border.top.style)
    dum.append(cell.border.bottom.style)
    dum.append(cell.border.right.style)
    dum.append(cell.border.left.style)


    #print('count_cell_border time is ',time.time()-start)

    return (dum.count('thin') if (dum.count('thin')>1) else None)


def count_cell_color(cell):
    try:
        temp = cell.fill.patternType
        if(temp=='solid'):
            val = 4
        else:
            val = None
    except:
        val = None

    return val


#THE BACKBONE OF THIS CODE, PARSE BORDER TO NUMPY ARRAY
def get_table_border(fname,sheetname):

    start = time.time()
    workbook = load_workbook(filename=fname,read_only=True)
    wb = workbook[sheetname]

    row = wb.max_row
    col = wb.max_column

    #Store to a list temporarily
    col_list = []
    for row in wb.rows:
        for cell in row:
            if(cell.border!=None):
                temp = count_cell_border(cell)
            else:
                temp = None
            col_list.append(temp)

    dframe = list_to_df(col_list,col)
    dframe = dframe.astype('Int64')

    return dframe


def get_table_color(fname,sheetname):

    start = time.time()
    workbook = load_workbook(filename=fname,read_only=True)
    wb = workbook[sheetname]

    row = wb.max_row
    col = wb.max_column

    #Store to a list temporarily
    col_list = []
    for row in wb.rows:
        for cell in row:

            if(cell.fill!=None):
                temp = count_cell_color(cell)
            else:
                temp = None

            col_list.append(temp)

    dframe = list_to_df(col_list,col)
    dframe = dframe.astype('Int64')

    return dframe


def get_table_style(fname,sheetname):

    start = time.time()
    workbook = load_workbook(filename=fname,read_only=True)
    wb = workbook[sheetname]

    row = wb.max_row
    col = wb.max_column

    #Store to a list temporarily
    col_list = []
    for row in wb.rows:
        for cell in row:

            if(cell.border!=None):
                style_border = [0 if count_cell_border(cell) is None else count_cell_border(cell)]
                color_border = [0 if count_cell_color(cell) is None else count_cell_color(cell)]
                temp = style_border[0] + color_border[0]
            else:
                temp = None

            col_list.append(temp)

    dframe = list_to_df(col_list,col)
    #dframe = dframe.astype(int)

    return dframe

##################################################################################

def get_border_by_row(df):
    start = time.time()
    val = []
    mark = True
    for row in df.index:
        for col in df.columns:
            cond = pd.isnull(df.iloc[row,col])

            if(cond ^ mark):
                if(cond):
                    val.append([row,col-1])
                else:
                    val.append([row,col])
            mark = cond

    #print('get_border_by_row time is ',time.time()-start)

    return val


def get_border_by_col(df):
    start = time.time()
    val = []
    mark = True
    for col in df.columns:
        for row in df.index:
            cond = pd.isnull(df.iloc[row,col])

            if(cond ^ mark):
                if(cond):
                    val.append([row-1,col])
                else:
                    val.append([row,col])
            mark = cond

    #print('get_border_by_col time is ',time.time()-start)
    return val



def find_edge_multitable(df):

    start = time.time()

    val_row = get_border_by_row(df)
    val_col = get_border_by_col(df)


    #print(val_row,end='')
    #print('\n\n')
    #print(val_col,end='')


    val_all = [value for value in val_row if value in val_col]
    #print(val_all)

    val_final=[]
    [val_final.append(x) for x in val_all if x not in val_final]
    #print(val_final,end='')



    #GET ROW
    tamp = sorted(val_final, key=operator.itemgetter(0))
    #print(len(tamp))


    #GET COLUMN
    temp = sorted(val_final, key=operator.itemgetter(1))
    #print(len(temp))



    lst = [item[0] for item in temp]
    res = []
    [res.append(x) for x in lst if x not in res]


    res_final = []
    for index in range(0,len(res),2):
        temp = res[index:index+2]
        res_final.append(min(temp))
        res_final.append(max(temp))

    cond = False
    ces = []
    for r in res:
        dummy = []
        idx = [x[0] for x in enumerate(tamp) if r == x[1][0]]
        for ids in idx:
            dummy.append(tamp[ids])
        dum = sorted(dummy, key=operator.itemgetter(1),reverse=cond)
        cond = not cond
        ces.append(dum[0][1])

    index = min(len(ces),len(res_final))

    if(index%2==0):
        pass
    else:
        index = index-1
    #print(index,len(ces),len(res_final))



    total_data = []
    for idx in range(0,index,2):
        val = (res_final[idx+1] - res_final[idx] + 1)*(ces[idx+1] - ces[idx]+1)
        total_data.append([val,idx,idx+1])

    final_data_index = sorted(total_data, key=operator.itemgetter(0), reverse=True)

    #print('find_edge_multitable time is ',time.time()-start)


    return res_final, ces, final_data_index

def auto_parse(filename, sheet, mode = 'border', debug=False):

    start = time.time()

    # Convert Table Style
    if(mode == 'all'):
        df = get_table_style(fname=filename, sheetname=sheet)
    elif(mode == 'border'):
        df = get_table_border(fname=filename, sheetname=sheet)
    elif(mode == 'color'):
        df = get_table_color(fname=filename, sheetname=sheet)
    else:
        print("f0# ~^4x $#%<")

    df = append_null_limiter(df)
    # df.fillna(0,inplace=True)
    # df.astype('Int32')

    if(debug):
        print('Execution time is ', time.time()-start)

    return df
